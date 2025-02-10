import sys
import json
import uuid
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
import qtawesome as qta
from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                            QHBoxLayout, QLabel, QPushButton, QComboBox, 
                            QTableWidget, QTableWidgetItem, QHeaderView, 
                            QMessageBox, QLineEdit, QDateEdit, QDialog,
                            QFormLayout, QTextEdit, QListWidget, QCheckBox,
                            QListWidgetItem, QMenu, QScrollArea, QFileDialog,
                            QStyle, QInputDialog)
from PySide6.QtCore import Qt, QDate
from PySide6.QtGui import QFont, QPalette, QColor, QIcon
from database import Database
from style_sheet import StyleSheet
from add_record_dialog import AddRecordDialog
from company_manager_dialog import CompanyManagerDialog
from vehicle_manager_dialog import VehicleManagerDialog
from wash_item_manager_dialog import WashItemManagerDialog
from company_dialog import CompanyDialog
from vehicle_dialog import VehicleDialog

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("電子紀錄系統")
        self.setStyleSheet(StyleSheet.MAIN_STYLE)
        self.setMinimumWidth(1200)
        self.setMinimumHeight(800)

        # 初始化資料
        self.data = {"companies": {}}
        self.database = Database()
        self.load_data()
        
        # 設置主要 widget 和布局
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout()
        layout.setSpacing(20)
        central_widget.setLayout(layout)

        # 上方控制區域
        top_layout = QHBoxLayout()
        
        # 公司選擇區域
        company_layout = QHBoxLayout()
        company_layout.setSpacing(2)  # 設置更小的間距
        company_label = QLabel("公司:")
        self.company_combo = QComboBox()
        self.company_combo.addItem("全部公司", "all")
        self.company_combo.setMinimumWidth(150)
        manage_company_btn = QPushButton()
        manage_company_btn.setIcon(qta.icon('fa5s.cog'))
        manage_company_btn.setToolTip("管理公司")
        manage_company_btn.setStyleSheet("""
            QPushButton {
                border: none;
                background-color: transparent;
                padding: 0px;
                margin: 0px;
            }
            QPushButton:hover {
                background-color: #f0f0f0;
            }
        """)
        manage_company_btn.setFixedSize(24, 24)  # 設置固定大小
        manage_company_btn.clicked.connect(self.manage_companies)
        company_layout.addWidget(company_label)
        company_layout.addWidget(self.company_combo)
        company_layout.addWidget(manage_company_btn)
        
        # 車輛選擇區域
        vehicle_layout = QHBoxLayout()
        vehicle_layout.setSpacing(2)  # 設置更小的間距
        vehicle_label = QLabel("車輛:")
        self.vehicle_combo = QComboBox()
        self.vehicle_combo.addItem("全部車輛", "all")
        self.vehicle_combo.setMinimumWidth(150)
        manage_vehicle_btn = QPushButton()
        manage_vehicle_btn.setIcon(qta.icon('fa5s.cog'))
        manage_vehicle_btn.setToolTip("管理車輛")
        manage_vehicle_btn.setStyleSheet("""
            QPushButton {
                border: none;
                background-color: transparent;
                padding: 0px;
                margin: 0px;
            }
            QPushButton:hover {
                background-color: #f0f0f0;
            }
        """)
        manage_vehicle_btn.setFixedSize(24, 24)  # 設置固定大小
        manage_vehicle_btn.clicked.connect(self.manage_vehicles)
        vehicle_layout.addWidget(vehicle_label)
        vehicle_layout.addWidget(self.vehicle_combo)
        vehicle_layout.addWidget(manage_vehicle_btn)

        # 將公司和車輛選擇區域添加到頂部佈局
        selection_layout = QHBoxLayout()
        selection_layout.addLayout(company_layout)
        selection_layout.addSpacing(20)  # 公司和車輛選擇之間的間距
        selection_layout.addLayout(vehicle_layout)
        top_layout.addLayout(selection_layout)
        top_layout.addStretch()

        # 新增紀錄和匯出按鈕
        buttons_layout = QHBoxLayout()
        add_record_btn = QPushButton("新增紀錄")
        add_record_btn.setMinimumWidth(150)
        add_record_btn.clicked.connect(self.add_record)
        export_btn = QPushButton("匯出篩選資料")
        export_btn.setMinimumWidth(150)
        export_btn.clicked.connect(self.export_excel)
        buttons_layout.addWidget(add_record_btn)
        buttons_layout.addWidget(export_btn)
        buttons_layout.setSpacing(10)
        buttons_layout.addStretch()
        top_layout.addLayout(buttons_layout)

        layout.addLayout(top_layout)

        # 搜尋區域
        search_layout = QHBoxLayout()
        
        # 日期範圍搜尋
        date_layout = QHBoxLayout()
        date_from_label = QLabel("日期從:")
        self.start_date = QDateEdit()
        self.start_date.setCalendarPopup(True)
        self.start_date.setDate(QDate.currentDate().addYears(-1))  # 預設為一年前
        date_to_label = QLabel("到:")
        self.end_date = QDateEdit()
        self.end_date.setCalendarPopup(True)
        self.end_date.setDate(QDate.currentDate())
        date_layout.addWidget(date_from_label)
        date_layout.addWidget(self.start_date)
        date_layout.addWidget(date_to_label)
        date_layout.addWidget(self.end_date)
        search_layout.addLayout(date_layout)

        # 搜尋欄位
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("搜尋服務項目、備註...")
        self.search_input.textChanged.connect(self.filter_records)
        search_layout.addWidget(self.search_input)

        # 清除搜尋按鈕
        clear_search_btn = QPushButton("清除搜尋")
        clear_search_btn.clicked.connect(self.clear_search)
        search_layout.addWidget(clear_search_btn)

        layout.addLayout(search_layout)

        # 表格
        self.table = QTableWidget()
        self.table.setColumnCount(9)  # 減少一欄
        self.table.setHorizontalHeaderLabels(["類型", "日期", "公司", "車牌號碼", "車輛種類", "服務項目", "備註", "金額總計", "操作"])
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)  # 禁止編輯
        header = self.table.horizontalHeader()
        
        # 設置表格標題可調整大小
        header.setSectionsMovable(True)  # 允許移動欄位
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)  # 允許使用者調整寬度
        
        # 設置預設寬度
        header.resizeSection(0, 80)  # 類型
        header.resizeSection(1, 100)  # 日期
        header.resizeSection(2, 100)  # 公司
        header.resizeSection(3, 150)  # 車牌號碼
        header.resizeSection(4, 100)  # 車輛種類
        header.resizeSection(5, 300)  # 服務項目
        header.resizeSection(7, 100)  # 金額總計
        header.resizeSection(8, 100)  # 操作
        
        # 設置最小寬度限制
        header.setMinimumSectionSize(60)
        
        # 設置服務項目和備註欄位可以自動延展
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.Stretch)  # 服務項目
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.Stretch)  # 備註
        self.table.verticalHeader().setVisible(False)
        self.table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                alternate-background-color: #f9f9f9;
            }
            QHeaderView::section {
                background-color: #e0e0e0;
                padding: 5px;
                border: none;
                border-right: 1px solid #ccc;
                border-bottom: 1px solid #ccc;
            }
            QTableWidget::item {
                padding: 5px;
                border-bottom: 1px solid #eee;
            }
        """)
        self.table.setAlternatingRowColors(True)
        self.table.setWordWrap(True)  # 啟用自動換行
        self.table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)  # 自動調整行高
        layout.addWidget(self.table)
        
        # 設置事件處理
        self.company_combo.currentIndexChanged.connect(self.update_vehicle_combo)
        self.company_combo.currentIndexChanged.connect(self.update_table)
        self.vehicle_combo.currentIndexChanged.connect(self.update_table)
        self.start_date.dateChanged.connect(self.update_table)
        self.end_date.dateChanged.connect(self.update_table)
        
        # 初始化下拉選單
        self.update_company_combo()
        self.update_vehicle_combo()
        
        # 更新表格
        self.update_table()

    def load_data(self):
        """從資料庫載入資料"""
        try:
            self.data = self.database.get_all_data()
            if not self.data:
                self.data = {"companies": {}}
        except Exception as e:
            QMessageBox.warning(self, "錯誤", f"載入資料時發生錯誤：{str(e)}")
            self.data = {"companies": {}}

    def save_data(self):
        """儲存資料到資料庫"""
        try:
            self.database.save_data(self.data)
        except Exception as e:
            QMessageBox.warning(self, "錯誤", f"儲存資料時發生錯誤：{str(e)}")

    def update_company_combo(self):
        """更新公司下拉選單"""
        self.company_combo.clear()
        self.company_combo.addItem("全部公司", "all")
        
        # 根據 sort_index 排序公司
        sorted_companies = sorted(
            self.data["companies"].items(),
            key=lambda x: x[1].get("sort_index", float('inf'))
        )
        
        for company_id, company_data in sorted_companies:
            self.company_combo.addItem(company_data["name"], company_id)

    def update_vehicle_combo(self):
        """更新車輛下拉選單"""
        self.vehicle_combo.clear()
        self.vehicle_combo.addItem("全部車輛", "all")
        company_id = self.company_combo.currentData()
        if company_id and company_id != "all":
            vehicles = self.data["companies"][company_id].get("vehicles", {})
            # 根據 sort_index 排序車輛
            sorted_vehicles = sorted(
                vehicles.items(),
                key=lambda x: x[1].get("sort_index", float('inf'))
            )
            for vehicle_id, vehicle_data in sorted_vehicles:
                self.vehicle_combo.addItem(f"{vehicle_data['plate']} ({vehicle_data['type']})", vehicle_id)

    def manage_companies(self):
        """管理公司"""
        dialog = CompanyManagerDialog(self, self.data)
        dialog.exec()
        self.update_company_combo()
        self.update_vehicle_combo()
        self.update_table()

    def manage_vehicles(self):
        """管理車輛"""
        company_id = self.company_combo.currentData()
        if company_id == "all":
            QMessageBox.warning(self, "警告", "請先選擇一個公司")
            return
        dialog = VehicleManagerDialog(self, company_id, self.data)
        dialog.exec()
        self.update_vehicle_combo()
        self.update_table()

    def add_record(self):
        """新增洗車紀錄"""
        dialog = AddRecordDialog(
            self,
            self.data,
            self.company_combo.currentData(),
            self.vehicle_combo.currentData()
        )
        if dialog.exec():
            record_data = dialog.get_record_data()
            company_id = record_data["company_id"]
            vehicle_id = record_data["vehicle_id"]
            
            if company_id == "all":
                QMessageBox.warning(self, "錯誤", "請選擇公司")
                return
                
            if not vehicle_id:
                QMessageBox.warning(self, "錯誤", "請選擇車輛")
                return
                
            # 添加紀錄到車輛資料中
            vehicle_data = self.data["companies"][company_id]["vehicles"][vehicle_id]
            if "records" not in vehicle_data:
                vehicle_data["records"] = []
            
            # 建立新的記錄
            new_record = {
                "date": record_data["date"],
                "items": record_data["items"],
                "remarks": record_data["remarks"],
                "payment_type": record_data["payment_type"]  # 添加應付/應收資訊
            }
            
            vehicle_data["records"].append(new_record)
            
            # 儲存並更新
            self.save_data()
            self.update_table()

    def search_records(self):
        """搜尋紀錄"""
        self.update_table()

    def update_table(self):
        """更新表格內容"""
        self.table.setRowCount(0)
        
        # 獲取當前選擇的公司和車輛
        selected_company_id = self.company_combo.currentData()
        selected_vehicle_id = self.vehicle_combo.currentData()
        
        # 獲取記錄
        records = []
        for company_id, company_data in self.data["companies"].items():
            # 如果選擇了特定公司，只顯示該公司的記錄
            if selected_company_id != "all" and company_id != selected_company_id:
                continue
                
            company_name = company_data["name"]
            for vehicle_id, vehicle_data in company_data.get("vehicles", {}).items():
                # 如果選擇了特定車輛，只顯示該車輛的記錄
                if selected_vehicle_id != "all" and vehicle_id != selected_vehicle_id:
                    continue
                    
                for record in vehicle_data.get("records", []):
                    record_with_info = record.copy()
                    record_with_info["company"] = company_name
                    record_with_info["company_id"] = company_id
                    record_with_info["vehicle_id"] = vehicle_id
                    record_with_info["vehicle"] = vehicle_data
                    records.append(record_with_info)
        
        # 填充表格
        for record in records:
            row = self.table.rowCount()
            self.table.insertRow(row)
            
            # 顯示應付/應收狀態
            payment_type = record.get("payment_type", "")
            payment_type_text = ""
            if payment_type == "payable":
                payment_type_text = "應付廠商"
            elif payment_type == "receivable":
                payment_type_text = "應收廠商"
            self.table.setItem(row, 0, QTableWidgetItem(payment_type_text))
            
            # 設置日期
            date_item = QTableWidgetItem(record["date"])
            self.table.setItem(row, 1, date_item)
            
            # 設置公司
            company_item = QTableWidgetItem(record["company"])
            company_item.setData(Qt.ItemDataRole.UserRole, record["company_id"])
            self.table.setItem(row, 2, company_item)
            
            # 設置車牌號碼
            vehicle_item = QTableWidgetItem(record["vehicle"]["plate"])
            vehicle_item.setData(Qt.ItemDataRole.UserRole, record["vehicle_id"])
            remarks = record["vehicle"].get("remarks", "")
            if remarks:
                vehicle_item.setToolTip(f"備註：{remarks}")
            self.table.setItem(row, 3, vehicle_item)
            
            # 確保表格項目可以顯示工具提示
            self.table.setMouseTracking(True)  # 啟用滑鼠追蹤
            self.table.viewport().setMouseTracking(True)
            
            # 設置車輛種類
            self.table.setItem(row, 4, QTableWidgetItem(record["vehicle"]["type"]))
            
            # 處理項目顯示和計算總金額
            items_text = []
            total_amount = 0
            for item in record["items"]:
                if isinstance(item, dict):
                    items_text.append(f"• {item['name']} - ${item['price']}")
                    total_amount += item['price']
                else:
                    items_text.append(f"• {item}")
            
            # 使用換行符號連接項目
            self.table.setItem(row, 5, QTableWidgetItem("\n".join(items_text)))
            
            # 設置備註
            self.table.setItem(row, 6, QTableWidgetItem(record.get("remarks", "")))
            
            # 設置金額總計
            total_amount_item = QTableWidgetItem(f"${total_amount:,}")
            total_amount_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table.setItem(row, 7, total_amount_item)
            
            # 添加刪除按鈕
            delete_btn = QPushButton("刪除")
            delete_btn.clicked.connect(lambda checked, r=row: self.delete_record(r))
            self.table.setCellWidget(row, 8, delete_btn)
            
            # 調整該行的高度以適應多行內容
            self.table.resizeRowToContents(row)

    def delete_record(self, row):
        """刪除記錄"""
        reply = QMessageBox.question(
            self,
            "確認刪除",
            "確定要刪除這筆紀錄嗎？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            try:
                # 從表格中獲取公司和車輛ID
                company_id = self.table.item(row, 2).data(Qt.ItemDataRole.UserRole)
                vehicle_id = self.table.item(row, 3).data(Qt.ItemDataRole.UserRole)
                date_str = self.table.item(row, 1).text()
                
                if not company_id or not vehicle_id:
                    QMessageBox.warning(self, "錯誤", "無法獲取記錄資訊")
                    return
                
                # 找到並刪除對應的記錄
                vehicle_records = self.data["companies"][company_id]["vehicles"][vehicle_id].get("records", [])
                for i, record in enumerate(vehicle_records):
                    if record["date"] == date_str:
                        # 從本地數據中刪除
                        vehicle_records.pop(i)
                        # 從 Firebase 中刪除
                        self.database.delete_record(company_id, vehicle_id, i)
                        break
                
                # 儲存並更新表格
                self.save_data()
                self.update_table()
                QMessageBox.information(self, "成功", "記錄已成功刪除！")
                
            except Exception as e:
                QMessageBox.warning(self, "錯誤", f"刪除記錄時發生錯誤：{str(e)}")

    def export_excel(self):
        """匯出資料到 Excel"""
        try:
            # 建立新的工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "洗車紀錄"
            
            # 寫入標題
            headers = ["類型", "日期", "公司", "車牌號碼", "車輛種類", "服務項目", "備註", "金額總計"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # 寫入資料
            excel_row = 2
            for table_row in range(self.table.rowCount()):
                # 獲取該行的服務項目
                service_items = self.table.item(table_row, 5).text().split('\n')
                first_item = True
                
                # 對每個服務項目創建一行
                for item in service_items:
                    # 如果是第一個項目，寫入所有欄位
                    if first_item:
                        for col in range(self.table.columnCount() - 1):  # 排除操作按鈕欄
                            if col == 5:  # 服務項目欄
                                ws.cell(row=excel_row, column=col + 1, value=item.strip())
                            else:
                                cell_item = self.table.item(table_row, col)
                                if cell_item:
                                    ws.cell(row=excel_row, column=col + 1, value=cell_item.text())
                        first_item = False
                    else:
                        # 如果不是第一個項目，只寫入服務項目
                        ws.cell(row=excel_row, column=6, value=item.strip())
                    excel_row += 1
                
                # 在每個記錄之間添加一個空行
                excel_row += 1
            
            # 調整欄寬
            for col in ws.columns:
                max_length = 0
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[col[0].column_letter].width = max_length + 2
            
            # 根據當前時間生成檔案名稱
            current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"洗車紀錄_{current_time}.xlsx"
            
            # 選擇儲存位置
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "儲存 Excel 檔案",
                default_filename,  # 預設檔案名稱
                "Excel 檔案 (*.xlsx)"
            )
            
            if file_path:
                if not file_path.endswith('.xlsx'):
                    file_path += '.xlsx'
                wb.save(file_path)
                QMessageBox.information(self, "成功", "資料已成功匯出！")
                
        except Exception as e:
            QMessageBox.warning(self, "錯誤", f"匯出資料時發生錯誤：{str(e)}")

    def save_wash_items(self, items):
        """儲存洗車項目"""
        try:
            self.database.save_wash_items(items)
        except Exception as e:
            QMessageBox.warning(self, "錯誤", f"儲存洗車項目時發生錯誤：{str(e)}")

    def filter_records(self):
        """根據搜尋條件過濾記錄"""
        search_text = self.search_input.text().lower().strip()
        
        # 獲取日期範圍
        start_date = self.start_date.date()
        end_date = self.end_date.date()

        for row in range(self.table.rowCount()):
            hide = False
            
            # 檢查日期
            date_str = self.table.item(row, 1).text()
            record_date = QDate.fromString(date_str, "yyyy-MM-dd")  # 先嘗試 yyyy-MM-dd 格式
            if not record_date.isValid():
                record_date = QDate.fromString(date_str, "yyyy/MM/dd")  # 如果無效，嘗試 yyyy/MM/dd 格式
            
            if record_date < start_date or record_date > end_date:
                hide = True            
            # 檢查搜尋文字
            if not hide and search_text:
                found = False
                
                # 檢查每個欄位
                for col in range(self.table.columnCount() - 1):  # 排除最後一欄（操作按鈕）
                    item = self.table.item(row, col)
                    if not item:
                        continue
                        
                    cell_text = item.text().lower()
                    
                    # 特別處理服務項目欄位
                    if col == 5:  # 服務項目欄位
                        # 分割每一行並檢查
                        lines = cell_text.split('\n')
                        for line in lines:
                            # 移除 "•" 和金額部分，但保留完整的項目名稱
                            service = line.replace('• ', '').split(' - $')[0].strip().lower()
                            if search_text in service:
                                found = True
                                break
                    else:
                        # 其他欄位直接檢查
                        if search_text in cell_text:
                            found = True
                            break
                
                if not found:
                    hide = True
            
            # 設置行的顯示/隱藏狀態
            self.table.setRowHidden(row, hide)

    def clear_search(self):
        """清除所有搜尋條件並重置顯示"""
        # 重置搜尋文字
        self.search_input.clear()
        
        # 重置日期範圍
        self.start_date.setDate(QDate.currentDate().addYears(-1))
        self.end_date.setDate(QDate.currentDate())
        
        # 重置公司和車輛選擇
        self.company_combo.setCurrentText("全部公司")
        self.vehicle_combo.setCurrentText("全部車輛")
        
        # 重新載入所有資料
        self.update_table()
        
        # 確保所有行都顯示
        for row in range(self.table.rowCount()):
            self.table.setRowHidden(row, False)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
